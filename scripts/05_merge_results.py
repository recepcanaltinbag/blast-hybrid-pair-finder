"""
STEP 5 — Merge everything into one final results table.

Joins:
  merged_summary.csv (query metadata)
  + blast_hits_filtered.tsv (BLAST results)
  + sra_metadata.tsv (all SRA runs + BioSample)
  + paired_runs.tsv  (long+short read pairs)

Output:
  results/final_results.tsv
  results/final_results.xlsx  — Sheet 1: all results
                                Sheet 2: long+short paired runs
"""

import os
import sys

import pandas as pd
from tqdm import tqdm

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import (
    SUMMARY_CSV, BLAST_HITS, SRA_META, PAIRED_RUNS,
    FINAL_TSV, FINAL_EXCEL,
    MIN_QUERY_COV, MIN_IDENTITY, MIN_PAIRS_PER_QUERY,
)

G   = "\033[92m"; Y = "\033[93m"; R = "\033[91m"
C   = "\033[96m"; W = "\033[97m"; DIM = "\033[2m"; RST = "\033[0m"; BOLD = "\033[1m"

def hr(char="─", width=62):
    return DIM + char * width + RST


def load_summary(csv_path: str) -> pd.DataFrame:
    df   = pd.read_csv(csv_path)
    mask = (df["CompTN"] == "Novel") & (df["ISproducts"].str.contains("IS26", na=False))
    df   = df[mask].copy()
    df["is26_type"] = df["ISproducts"].fillna("").str.strip().str.replace(" ", "_")
    df["ant_safe"]  = df["Antproducts_names"].fillna("no_AMR").str.replace(";", "|")
    df["query_id"]  = (df["SRAID"].astype(str) + "|" +
                       df["CycleID"].astype(str) + "|" +
                       df["is26_type"] + "|" +
                       df["ant_safe"])
    keep = ["query_id", "SRAID", "CycleID", "is26_type",
            "Antproducts_names", "NumAnt", "NumIS",
            "score0", "score1", "score2"]
    return df[keep].rename(columns={
        "SRAID"             : "source_sra",
        "CycleID"           : "cycle_id",
        "Antproducts_names" : "resistance_genes",
        "NumAnt"            : "n_resistance_genes",
        "NumIS"             : "n_is_elements",
        "score0"            : "pipeline_score0",
        "score1"            : "pipeline_score1",
        "score2"            : "pipeline_score2",
    })


def style_sheet(ws, header_color: str, light_color: str, header_font):
    from openpyxl.styles import PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    hfill = PatternFill(fill_type="solid", fgColor=header_color)
    lfill = PatternFill(fill_type="solid", fgColor=light_color)
    for cell in ws[1]:
        cell.fill      = hfill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        if row_idx % 2 == 0:
            for cell in row:
                cell.fill = lfill
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)
    ws.freeze_panes = "A2"


def main():
    print()
    print(hr("═"))
    print(f"  {BOLD}{C}STEP 05 — MERGE & EXPORT{RST}")
    print(hr("═"))

    steps = [
        ("Loading summary CSV",         SUMMARY_CSV),
        ("Loading BLAST hits",          BLAST_HITS),
        ("Loading SRA metadata",        SRA_META),
        ("Loading paired runs",         PAIRED_RUNS),
    ]
    for label, path in steps:
        exists = os.path.exists(path)
        icon   = f"{G}✓{RST}" if exists else f"{R}✗{RST}"
        print(f"  {icon}  {label:<30} {DIM}{os.path.basename(path)}{RST}")
        if not exists:
            print(f"\n  {R}ERROR: {path} not found.{RST}")
            sys.exit(1)

    print(hr())

    with tqdm(total=6, bar_format=f"  {C}{{l_bar}}{{bar}}{RST}{DIM}| {{n_fmt}}/{{total_fmt}}{RST}",
              ncols=68) as pbar:

        pbar.set_description(f"  {C}Reading files{RST}")
        summary  = load_summary(SUMMARY_CSV)
        blast    = pd.read_csv(BLAST_HITS,  sep="\t")
        sra_meta = pd.read_csv(SRA_META,    sep="\t")
        pairs_df = pd.read_csv(PAIRED_RUNS, sep="\t")
        pbar.update(1)

        pbar.set_description(f"  {C}Joining tables{RST}")
        merged = summary.merge(blast, on="query_id", how="left")
        merged = merged.merge(sra_meta, on="hit_accession", how="left")
        merged["organism"] = (merged.get("organism_sra", pd.Series(dtype=str))
                                    .fillna(merged.get("organism_blast", pd.Series(dtype=str))))
        merged.drop(columns=["organism_sra", "organism_blast"], inplace=True, errors="ignore")

        # Pull BioSample metadata (country, host, etc.) from paired_runs into merged
        # (these are only populated in make_pairs via get_biosample_meta)
        bs_meta_cols = ["hit_accession", "country", "isolation_source", "host",
                        "collection_date", "strain", "biosample"]
        bs_meta_cols_avail = [c for c in bs_meta_cols if c in pairs_df.columns]
        if len(bs_meta_cols_avail) > 1:
            bs_meta = (pairs_df[bs_meta_cols_avail]
                       .drop_duplicates(subset=["hit_accession"])
                       .rename(columns={c: c + "_pair" if c != "hit_accession" else c
                                        for c in bs_meta_cols_avail}))
            # Rename back (no suffix clash since sra_meta doesn't have these cols)
            rename_map = {c + "_pair": c for c in bs_meta_cols_avail if c != "hit_accession"}
            bs_meta = bs_meta.rename(columns=rename_map)
            for col in [c for c in bs_meta_cols_avail if c != "hit_accession"]:
                if col not in merged.columns:
                    merged = merged.merge(bs_meta[["hit_accession", col]], on="hit_accession", how="left")

        # Enrich paired_runs with query/cycle info
        # query_id → hit_accession mapping from blast hits
        query_map = (blast[["query_id", "hit_accession"]]
                     .drop_duplicates()
                     .merge(summary[["query_id", "source_sra", "cycle_id",
                                     "is26_type", "resistance_genes"]],
                            on="query_id", how="left"))
        # Group: one accession may match multiple queries
        query_agg = (query_map.groupby("hit_accession")
                              .agg(
                                  query_ids       =("query_id",        lambda x: "|".join(x.dropna().unique())),
                                  source_sras     =("source_sra",      lambda x: "|".join(x.dropna().unique())),
                                  cycle_ids       =("cycle_id",        lambda x: "|".join(x.dropna().astype(str).unique())),
                                  is26_types      =("is26_type",       lambda x: "|".join(x.dropna().unique())),
                                  resistance_genes=("resistance_genes", lambda x: "|".join(x.dropna().unique())),
                              )
                              .reset_index())
        pairs_df = pairs_df.merge(query_agg, on="hit_accession", how="left")

        # Reorder paired_runs columns: put query info first
        pair_col_order = [
            "hit_accession", "query_ids", "source_sras", "cycle_ids",
            "is26_types", "resistance_genes",
            "biosample", "bioproject", "organism_sra", "organism_blast",
            "long_read_run", "long_read_platform", "long_read_strategy",
            "long_read_bases_Mbp", "long_read_spots",
            "short_read_run", "short_read_platform", "short_read_strategy",
            "short_read_layout", "short_read_bases_Mbp", "short_read_spots",
            "n_long_runs", "n_short_runs",
            "isolation_source", "host", "country", "collection_date", "strain",
        ]
        pair_col_order = [c for c in pair_col_order if c in pairs_df.columns]
        remaining_pair = [c for c in pairs_df.columns if c not in pair_col_order]
        pairs_df = pairs_df[pair_col_order + remaining_pair]
        pbar.update(1)

        pbar.set_description(f"  {C}Ordering columns{RST}")
        col_order = [
            "query_id", "source_sra", "cycle_id", "is26_type",
            "resistance_genes", "n_resistance_genes", "n_is_elements",
            "pipeline_score0", "pipeline_score1", "pipeline_score2",
            "hit_accession", "hit_title", "organism",
            "identity_pct", "query_cov_pct", "evalue", "bitscore",
            "align_len", "query_len",
            "sra_run", "platform", "library_strategy", "library_layout",
            "spots", "bases_Mbp",
            "bioproject", "biosample",
            "isolation_source", "host", "country",
            "collection_date", "strain", "submitter_org",
        ]
        col_order = [c for c in col_order if c in merged.columns]
        remaining = [c for c in merged.columns if c not in col_order]
        merged    = merged[col_order + remaining]
        pbar.update(1)

        pbar.set_description(f"  {C}Saving TSV{RST}")
        os.makedirs(os.path.dirname(FINAL_TSV), exist_ok=True)
        merged.to_csv(FINAL_TSV, sep="\t", index=False)
        pbar.update(1)

        pbar.set_description(f"  {C}Building Excel{RST}")
        try:
            from openpyxl.styles import Font
            hfont = Font(color="FFFFFF", bold=True)
            with pd.ExcelWriter(FINAL_EXCEL, engine="openpyxl") as writer:
                merged.to_excel(writer, index=False, sheet_name="Query_Results")
                pairs_df.to_excel(writer, index=False, sheet_name="Paired_Runs")
                style_sheet(writer.sheets["Query_Results"],   "1F4E79", "DCE6F1", hfont)
                style_sheet(writer.sheets["Paired_Runs"],    "1E6B3C", "D6EFD8", hfont)
        except ImportError:
            tqdm.write(f"  {Y}[!] openpyxl not installed — skipping Excel.{RST}")
        pbar.update(1)

        pbar.set_description(f"  {C}Done{RST}")
        pbar.update(1)

    # ── Summary ────────────────────────────────────────────────────────────
    has_hits    = merged["hit_accession"].notna().sum()
    has_hits_pct = has_hits / len(merged) * 100 if len(merged) else 0
    uniq_org    = merged["organism"].dropna().nunique()
    uniq_cntry  = merged["country"].dropna().nunique() if "country" in merged.columns else 0

    pair_counts  = pairs_df.groupby("hit_accession").size() if not pairs_df.empty and "hit_accession" in pairs_df.columns else pd.Series(dtype=int)
    meeting      = (pair_counts >= MIN_PAIRS_PER_QUERY).sum() if not pair_counts.empty else 0

    print()
    print(hr("═"))
    print(f"  {BOLD}{G}MERGE COMPLETE{RST}")
    print(hr("═"))
    print(f"  {W}Total rows              :{RST} {len(merged)}")
    print(f"  {W}Unique queries          :{RST} {merged['query_id'].nunique()}")
    print(f"  {W}Queries with BLAST hits :{RST} {G}{has_hits}{RST} ({has_hits_pct:.0f}%)")

    bar = "█" * int(has_hits_pct / 5)
    print(f"  {G}{bar:<20}{RST} {has_hits_pct:.0f}%")

    print(hr())
    print(f"  {W}Unique organisms        :{RST} {uniq_org}")
    print(f"  {W}Unique countries        :{RST} {uniq_cntry}")
    print(f"  {W}Long+short pairs total  :{RST} {len(pairs_df)}")
    print(f"  {W}Accs with ≥{MIN_PAIRS_PER_QUERY} pairs     :{RST} {G}{meeting}{RST} / {len(pair_counts)}")

    if not pairs_df.empty and "long_read_platform" in pairs_df.columns:
        print(hr())
        print(f"  {W}Long-read platforms in pairs:{RST}")
        for plat, cnt in pairs_df["long_read_platform"].value_counts().items():
            b = "█" * min(int(cnt / pairs_df["long_read_platform"].value_counts().max() * 20), 20)
            print(f"    {C}{b:<20}{RST} {cnt:>4}  {DIM}{plat}{RST}")

    if "resistance_genes" in merged.columns:
        print(hr())
        all_genes = (merged["resistance_genes"].dropna()
                            .str.split(";").explode().str.strip()
                            .value_counts().head(5))
        print(f"  {W}Top 5 resistance genes:{RST}")
        for gene, cnt in all_genes.items():
            b = "█" * min(int(cnt / all_genes.max() * 20), 20)
            print(f"    {Y}{b:<20}{RST} {cnt:>4}  {DIM}{gene}{RST}")

    print(hr("═"))
    print(f"\n  {G}→ TSV  :{RST} {FINAL_TSV}")
    print(f"  {G}→ XLSX :{RST} {FINAL_EXCEL}")
    print(f"  {G}→ Next :{RST} python3 scripts/06_visualize.py\n")


if __name__ == "__main__":
    main()
